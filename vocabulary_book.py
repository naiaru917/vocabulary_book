#!/usr/bin/env python
# coding: utf-8

# In[1]:


#モジュールを取り込む
import openpyxl
#ファイルを指定する
wb = openpyxl.load_workbook(r'C:\Users\tfust\Documents\excel_test.xlsx')
#シートを指定する
ws = wb.worksheets[1]  #一番最初のシートは0
#列データをリスト化する
listQ = [ x.value for x in ws['A'] ]  #B列(問題)
listA = [ y.value for y in ws['B'] ]  #A列(回答)

#リストの表示（確認用）
#print(list1)
#print(list2)

cnt = 0  #問題数を数える
score = 0  #正解数を数える

for x in listQ:
    cnt += 1
    print(f"第{cnt}問:",x)
    ans = input("答えを入力してください")
    if ans == listA[cnt - 1]:
        print ('\033[31m' +"正解"+ '\033[0m')    #文字を赤色にする
        score += 1
    else:
        print("不正解" )
        print("正解は",'\033[31m' + listA[cnt-1] + '\033[0m')
    print()

#print("正答率:",score,"/",cnt)
# In[4]:


import random

#重複なしのランダムな数列
def rand_ints_nodup(a, b, k):
  ns = []
  while len(ns) < k:
    n = random.randint(a, b)
    if not n in ns:
      ns.append(n)
  return ns

L = len(listQ)   #リストの個数を求める
listRnd = rand_ints_nodup(1, L, 10)   #1~Lからランダムで10個
print(listRnd)


# In[5]:


cnt = 0
score = 0
for x in listRnd:
    cnt += 1
    print(f"第{cnt}問:",listQ[x-1])
    ans = input("答えを入力してください")
    if ans == listA[x-1]:
        print ('\033[31m' +"正解"+ '\033[0m')    #文字を赤色にする
        score += 1
    else:
        print("不正解" )
        print("正解は",'\033[31m' + listA[x-1] + '\033[0m')
    print()

print("正答率:",score,"/",cnt)


# In[ ]:




